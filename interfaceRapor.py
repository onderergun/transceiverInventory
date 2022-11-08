import csv
from jsonrpclib import Server
import ssl
from getpass import getpass
import time
import argparse

import openpyxl
from openpyxl.styles import Font

ssl._create_default_https_context = ssl._create_unverified_context

d1 = time.strftime("%Y_%m_%d_%H_%M_%S", time.gmtime())

parser = argparse.ArgumentParser()
parser.add_argument('--username', required=True)
parser.add_argument('--inventoryfile', required=True)
args = parser.parse_args()
switchuser = args.username
inventoryfilename=args.inventoryfile
switchpass = getpass()
filename= "Interface_report_" + d1 + ".xlsx"
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hostname'
sheet['A1'].font = Font(size=14, bold=True)
sheet.column_dimensions['A'].width = 20
sheet['B1'] = 'Model'
sheet['B1'].font = Font(size=14, bold=True)
sheet.column_dimensions['B'].width = 20
sheet['C1'] = 'Serial Number'
sheet['C1'].font = Font(size=14, bold=True)
sheet.column_dimensions['C'].width = 20
sheet['D1'] = 'Interface'
sheet['D1'].font = Font(size=14, bold=True)
sheet.column_dimensions['D'].width = 15
sheet['E1'] = 'Link Status'
sheet['E1'].font = Font(size=14, bold=True)
sheet.column_dimensions['E'].width = 20
sheet['F1'] = 'Interface Type'
sheet['F1'].font = Font(size=14, bold=True)
sheet.column_dimensions['F'].width = 20

k=0
with open(inventoryfilename) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    for iter, row in enumerate(csv_reader):
        if iter == 0:
            for num, column in enumerate(row):
                if column == "IP Address" or column == "IPAddress":
                    hostIndex = num
        else:
            ssh_host=row[hostIndex]
            hostname=row[0]
            serialnumber=row[7]
            SKU=row[2]
            if "vEOS" not in SKU and "OOB" not in hostname:
                print (hostname)
                urlString = "https://{}:{}@{}/command-api".format(switchuser, switchpass, ssh_host)
                switchReq = Server(urlString)
                response = switchReq.runCmds( 1, ["enable", "show interfaces status"])
                intstat = response[1]["interfaceStatuses"]
                for interface in intstat:
                    if "Ethernet" in interface:
                        sheet.cell(row=k+2,column=1).value = hostname
                        sheet.cell(row=k+2,column=2).value = SKU
                        sheet.cell(row=k+2,column=3).value = serialnumber
                        sheet.cell(row=k+2,column=4).value = interface
                        sheet.cell(row=k+2,column=5).value = intstat[interface]["linkStatus"]
                        sheet.cell(row=k+2,column=6).value = intstat[interface]["interfaceType"]
                        k+=1

wb.save(filename)


