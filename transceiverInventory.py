import csv
from getpass import getpass
import os
import argparse
import openpyxl
from openpyxl.styles import Font
import time
import requests
import json
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

parser = argparse.ArgumentParser()
parser.add_argument('--username', required=True)
parser.add_argument('--inventoryname', required=True)


args = parser.parse_args()
switchuser = args.username
inventory = args.inventoryname
switchpass = getpass()

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hostname'
sheet['A1'].font = Font(size=14, bold=True)
sheet.column_dimensions['A'].width = 30
sheet['B1'] = 'XCVR Model Name'
sheet['B1'].font = Font(size=14, bold=True)
sheet.column_dimensions['B'].width = 20
sheet['C1'] = 'Serial Number'
sheet['C1'].font = Font(size=14, bold=True)
sheet.column_dimensions['C'].width = 20
rownum = 3

with open(inventory) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    for iter,row in enumerate(csv_reader):
        if iter == 0 :
            for num,column in enumerate(row):
                if column == "IP Address" or column == "IPAddress":
                    hostIndex=num
        else:
            ssh_host=row[hostIndex]
            print (ssh_host)
            hostname=row[0]
            print (hostname)
            mgmtIP=row[5]
            urlString = "https://{}:{}@{}/command-api".format(switchuser, switchpass, ssh_host)
            data = json.dumps({
                'jsonrpc': '2.0',
                'method': 'runCmds',
                'params': {
                    'format': 'json',
                    'timestamps': False,
                    'autoComplete': False,
                    'expandAliases': False,
                    'cmds': [
                        'show inventory'
                    ],
                    'version': 1
                },
                'id': 'EapiExplorer-1'
            })
            response = requests.post(urlString, data=data, verify=False)
            output = json.loads(response.text)
            responsePr = output["result"][0]["xcvrSlots"]
            
            for key in responsePr:
                if responsePr[key]["serialNum"] != "":
                    sheet.cell(row=rownum,column=1).value = hostname
                    sheet.cell(row=rownum,column=2).value=responsePr[key]["modelName"]
                    sheet.cell(row=rownum,column=3).value=responsePr[key]["serialNum"]
                    rownum +=1

d1 = time.strftime("%Y_%m_%d_%H_%M_%S", time.gmtime())
filename = 'XCVR_Inventory_'+ d1 + '.xlsx'
wb.save(filename)

 